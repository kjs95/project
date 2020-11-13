import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
#크롤링 코드 입니다.
RESULT_PATH = 'C:/Users/user/Documents/news_crawling/'


def gain(url):
    info = []
    breq = requests.get("https://news.naver.com" + url)
    bs = BeautifulSoup(breq.content, 'html.parser')

    title = bs.select('h3#articleTitle')[0].text
    info.append(title)

    date = bs.select('.t11')[0].get_text()[:11]
    info.append(date)

    content = bs.select('#articleBodyContents')[0].get_text().replace('\n', " ")
    contents = content.replace("// flash 오류를 우회하기 위한 함수 추가 function _flash_removeCallback() {}", "")
    info.append(contents.strip())

    info.append(url)

    company = bs.select('#footer address')[0].a.get_text()
    info.append(company)

    return info


def crawler():
    # write_wb = Workbook()
    # write_ws = write_wb.active
    # write_ws.append(['날짜','회사','제목','내용','url'])
    s_date = [i for i in range(20191001, 20191032)]  # input('날짜 입력(20200505) : ')

    for j in num:
        write_wb = Workbook()
        write_ws = write_wb.active
        write_ws.append(['날짜', '회사', '제목', '내용', 'url'])

        for i in s_date:
            url = "https://news.naver.com/main/ranking/popularDay.nhn?rankingType=popular_day&sectionId={}&date=".format(
                j) + str(i)

            req = requests.get(url)
            soup = BeautifulSoup(req.content, 'html.parser')

            for urls in soup.select(".ranking_headline > a"):
                try:

                    if urls["href"]:
                        news_detail = gain(urls["href"])

                        write_ws.append(
                            [news_detail[1], news_detail[4], news_detail[0], news_detail[2], news_detail[3]])

                except Exception as e:
                    print(e)
                    continue

        write_wb.save('C:/Users/user/Documents/news_crawling/{}_{}_{}.xlsx'.format(s_date[0], s_date[-1], j))


num = [i for i in range(100, 106)]  # input('100 : 정치, 101 : 경제, 102 : 사회, 103 : 생활/문화, 104 : 세계, 105 : 과학/it 숫자 입력해 주세요 : ')
crawler()
# import requests
# from bs4 import BeautifulSoup
# from openpyxl import Workbook
#
# RESULT_PATH = 'C:/Users/user/Documents/news_crawling/'
#
#
#
# def gain(url):
#     info = []
#     breq = requests.get("https://news.naver.com"+url)
#     bs = BeautifulSoup(breq.content, 'html.parser')
#
#     title = bs.select('h3#articleTitle')[0].text
#     info.append(title)
#
#     date = bs.select('.t11')[0].get_text()[:11]
#     info.append(date)
#
#     content = bs.select('#articleBodyContents')[0].get_text().replace('\n', " ")
#     contents = content.replace("// flash 오류를 우회하기 위한 함수 추가 function _flash_removeCallback() {}", "")
#     info.append(contents.strip())
#
#     info.append(url)
#
#     company = bs.select('#footer address')[0].a.get_text()
#     info.append(company)
#
#     return info
#
#
# def crawler():
#
#
#
#     write_wb = Workbook()
#     write_ws = write_wb.active
#     write_ws.append(['날짜','회사','제목','내용','url'])
#     s_date = [i for i in range(20200401,20200431)] #input('날짜 입력(20200505) : ')
#
#
#
#     for i in s_date:
#         url = "https://news.naver.com/main/ranking/popularDay.nhn?rankingType=popular_day&sectionId={}&date=".format(num) + str(i)
#
#         req = requests.get(url)
#         soup = BeautifulSoup(req.content, 'html.parser')
#
#
#         for urls in soup.select(".ranking_headline > a"):
#             try:
#
#                 if urls["href"]:
#                     news_detail = gain(urls["href"])
#
#                     write_ws.append([news_detail[1],news_detail[4],news_detail[0],news_detail[2],news_detail[3]])
#
#             except Exception as e:
#                 print(e)
#                 continue
#
#
#
#     write_wb.save('C:/Users/user/Documents/news_crawling/{}_{}._{}xlsx'.format(s_date[0],s_date[-1],num))
#
#
#
#
#
#
#
#
# num = [i for i in range(100,106)]#input('100 : 정치, 101 : 경제, 102 : 사회, 103 : 생활/문화, 104 : 세계, 105 : 과학/it 숫자 입력해 주세요 : ')
# crawler()